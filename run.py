import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
import openpyxl


tf.set_random_seed(777)

# train Parameters
seq_length = 7
data_dim = 8
hidden_dim = 5
output_dim = 1
num_stacked_layers = 3
learning_rate = 0.01
num_epoch = 300
check_step = 10

def MinMaxScaler(data):
	numerator = data - np.min(data, 0)
	denominator = np.max(data, 0) - np.min(data, 0)
	return numerator / (denominator+ 1e-7)

def reverse(data, origin):
	min = np.min(origin, 0)
	max = np.max(origin, 0)
	return data * (max[0] - min[0] + 1e-7) + min[0]

def DataGen(name):
	xy = np.load("./data/" + name + ".npy")
	xy = MinMaxScaler(xy)
	x = xy
	y = xy[:, 0]
	dataX = []
	dataY = []
	for i in range(0, len(y) - seq_length):
		_x = x[i: i + seq_length]
		_y = y[i + seq_length]
		dataX.append(_x)
		dataY.append(_y)
	return dataX, dataY


def lstm_cell(ReLu = False):
	if ReLu:
		return tf.contrib.rnn.BasicLSTMCell(num_units=hidden_dim, state_is_tuple=True, activation=tf.nn.relu)
	return tf.contrib.rnn.BasicLSTMCell(num_units=hidden_dim, state_is_tuple=True, activation=tf.tanh)


def rnn_cell(ReLu = False):
	if ReLu:
		return tf.contrib.rnn.BasicRNNCell(num_units=hidden_dim, state_is_tuple=True, activation=tf.nn.relu)
	return tf.contrib.rnn.BasicRNNCell(num_units=hidden_dim, state_is_tuple=True, activation=tf.tanh)


trainData = DataGen("train")
testData = DataGen("test")

trainX, trainY = trainData[0], np.reshape(trainData[1], (-1, 1))
testX, testY = testData[0], np.reshape(testData[1], (-1, 1))

X = tf.placeholder(tf.float32, [None, seq_length, data_dim])
Y = tf.placeholder(tf.float32, [None, 1])

multi_cells = tf.contrib.rnn.MultiRNNCell([lstm_cell() for _ in range(num_stacked_layers)], state_is_tuple=True)
outputs, _ = tf.nn.dynamic_rnn(multi_cells, X, dtype=tf.float32)
Y_pred = tf.contrib.layers.fully_connected(outputs[:, -1], output_dim, activation_fn=None)

loss = tf.reduce_sum(tf.square(Y_pred - Y))

optimizer = tf.train.AdamOptimizer(learning_rate)
train = optimizer.minimize(loss)

targets = tf.placeholder(tf.float32, [None, 1])
predictions = tf.placeholder(tf.float32, [None, 1])
rmse = tf.sqrt(tf.reduce_mean(tf.square(targets - predictions)))


with tf.Session() as sess:
	sess.run(tf.global_variables_initializer())

	for epoch in range(1, num_epoch + 1):
		_, step_loss = sess.run([train, loss], feed_dict = {X: trainX, Y: trainY})

		if epoch % check_step ==0:
			print("[step: {}] loss: {}".format(epoch, step_loss))

	test_predict = sess.run(Y_pred, feed_dict = {X: testX})
	rmse_val = sess.run(rmse, feed_dict={targets: testY, predictions: test_predict})
	print("RMSE: {}".format(rmse_val))

	# plt.plot(testY)
	# plt.plot(test_predict)
	# plt.xlabel("Time Period")
	# plt.ylabel("Stock Price")
	# plt.show()

	tx = np.load("./data/test.npy")
	testdataY = reverse(testY, tx)

	sum = 0.
	result = []
	for i in range(len(testX)):
		lastX = np.reshape(testX[i], (-1, seq_length, data_dim))
		lastPredict = sess.run(Y_pred, feed_dict={X: lastX})
		lastPredict = reverse(lastPredict, tx)
		aa = abs(testdataY[i, 0] - lastPredict[0, 0]) / testdataY[i] * 100
		sum = sum + aa
		# print(lastPredict[0,0],testdataY[i],100-aa)

		result.append(lastPredict[0, 0])
	print("average: ", 100 - sum / len(testX))

	wr = openpyxl.load_workbook('./score.xlsx')
	wrr = wr.active

	wrr.cell(row=1, column=1).value = "Predict"
	wrr.cell(row=1, column=2).value = "Truth"
	for i in range(1, len(testX) + 1):
		wrr.cell(row=i + 1, column=1).value = result[i - 1]
		wrr.cell(row=i + 1, column=2).value = testdataY[i - 1, 0]

	wr.save("./score.xlsx")
	wr.close()

	plt.plot(result)
	plt.plot(testdataY)
	plt.show()

