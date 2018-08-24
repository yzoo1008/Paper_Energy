import numpy as np
import openpyxl
import os


if not "data" in os.listdir("./"):
	os.mkdir("data")

wb = openpyxl.load_workbook('./data.xlsx', data_only=True)
sheets = wb.worksheets

for i in range(1, 6):
	sheet = sheets[i]
	data = []
	row_c = sheet.max_row
	col_c = sheet.max_column
	for row in range(2, row_c + 1):
		row_data = []
		for col in range(0, 7):
			row_data.append(sheet[row][col].value)
		data.append(row_data)
	np.save("./data/" + str(sheet.title) + ".npy", data)


# for i in range(2014, 2019):
# 	array = np.load("./data/" + str(i) + ".npy")
# 	print(np.shape(array))
