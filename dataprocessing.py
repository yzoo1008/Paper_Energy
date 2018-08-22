import numpy as np
import openpyxl
import os


train = dict()
test = dict()
f_test = dict()


for year in range(2014, 2015):
	wb = openpyxl.load_workbook('./전력/' + str(year) + '.xlsx', data_only=True)
	sheet = wb.worksheets[0]
	row_c = sheet.max_row
	col_c = sheet.max_column
	for row in range(row_c, 1, -1):
		y = sheet[row][0].value
		m = sheet[row][1].value
		d = sheet[row][2].value
		date = y + m + d
		train[date] = []
		train[date].append(sheet[row][5].value)

for year in range(2014, 2015):
	wb = openpyxl.load_workbook('./기후/' + str(year) + '.xlsx', data_only=True)
	sheet = wb.worksheets[0]
	row_c = sheet.max_row
	col_c = sheet.max_column
	for row in range(2, row_c + 1):
		if str(sheet[row][0].value) == "108":
			ymd = str(sheet[row][1].value).split(" ")[0]
			ymd_ = ymd.split("-")
			y = ymd_[0]
			m = ymd_[1]
			d = ymd_[2]
			date = y + m + d
			for col in range(2, 9):
				if sheet[row][col].value:
					train[date].append(sheet[row][col].value)
				else:
					train[date].append(0)


for year in range(2018, 2019):
	wb = openpyxl.load_workbook('./전력/' + str(year) + '.xlsx', data_only=True)
	sheet = wb.worksheets[0]
	row_c = sheet.max_row
	col_c = sheet.max_column
	for row in range(row_c, 1, -1):
		y = sheet[row][0].value
		m = sheet[row][1].value
		d = sheet[row][2].value
		date = y + m + d
		test[date] = []
		test[date].append(sheet[row][5].value)

for year in range(2018, 2019):
	wb = openpyxl.load_workbook('./기후/' + str(year) + '.xlsx', data_only=True)
	sheet = wb.worksheets[0]
	row_c = sheet.max_row
	col_c = sheet.max_column
	for row in range(2, row_c + 1):
		if str(sheet[row][0].value) == "108":
			ymd = str(sheet[row][1].value).split(" ")[0]
			ymd_ = ymd.split("-")
			y = ymd_[0]
			m = ymd_[1]
			d = ymd_[2]
			date = y + m + d
			for col in range(2, 9):
				if sheet[row][col].value:
					test[date].append(sheet[row][col].value)
				else:
					test[date].append(0)


for year in range(2015, 2016):
	wb = openpyxl.load_workbook('./전력/' + str(year) + '.xlsx', data_only=True)
	sheet = wb.worksheets[0]
	row_c = sheet.max_row
	col_c = sheet.max_column
	for row in range(row_c, 1, -1):
		y = sheet[row][0].value
		m = sheet[row][1].value
		d = sheet[row][2].value
		date = y + m + d
		f_test[date] = []
		f_test[date].append(sheet[row][5].value)

for year in range(2015, 2016):
	wb = openpyxl.load_workbook('./기후/' + str(year) + '.xlsx', data_only=True)
	sheet = wb.worksheets[0]
	row_c = sheet.max_row
	col_c = sheet.max_column
	for row in range(2, row_c + 1):
		if str(sheet[row][0].value) == "108":
			ymd = str(sheet[row][1].value).split(" ")[0]
			ymd_ = ymd.split("-")
			y = ymd_[0]
			m = ymd_[1]
			d = ymd_[2]
			date = y + m + d
			for col in range(2, 9):
				if sheet[row][col].value:
					f_test[date].append(sheet[row][col].value)
				else:
					f_test[date].append(0)


key_train = sorted(train)
key_test = sorted(test)
key_f_test = sorted(f_test)
train_npy = []
test_npy = []
f_test_npy = []

for key in key_train:
	print(key, train[key])
	train_npy.append(train[key])

for key in key_test:
	print(key, test[key])
	test_npy.append(test[key])

for key in key_f_test:
	print(key, f_test[key])
	f_test_npy.append(f_test[key])

if not "data" in os.listdir("./"):
	os.mkdir("data")

np.save("./data/train.npy", train_npy)
np.save("./data/test.npy", test_npy)
np.save("./data/f_test.npy", f_test_npy)
